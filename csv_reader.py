import openpyxl
from engine.models import Industries
# Load the xlsx file
workbook = openpyxl.load_workbook('industries.xlsx')
# workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook.active

# Select the worksheet
# worksheet = workbook.active
# col1 = worksheet['A1'].value
# col2 = worksheet['B1'].value
# print(col1, col2)
#
# # # Iterate through rows and columns
# for row in worksheet['A1:C2']:
#     for cell in row:
#         print(cell.value)
#     for cell in row:
#         print(cell.value)

for row in sheet.iter_rows(min_row=2, values_only=True):
    try:
        if row[0] and row[1]:
            Industries.objects.create(title=row[0], slogan=row[1])
    except Exception as e:
        continue



